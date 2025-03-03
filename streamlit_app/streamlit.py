import cv2
from ultralytics import YOLO
import streamlit as st
import altair as alt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import base64
import os
import sys

try:
    from streamlit_webrtc import webrtc_streamer, VideoTransformerBase
except ImportError as e:
    st.error(f"Failed to import streamlit_webrtc: {e}")
    st.stop()

st.title("ElektroXen App")

# Get the path to the model file
if getattr(sys, 'frozen', False):
    # If the application is run as a bundle, use this path for the model file
    model_path = os.path.join(sys._MEIPASS, r'streamlit_app/best_F3.pt')
else:
    # If running in a normal environment, use the usual path
    model_path = r'streamlit_app/best_F3.pt'

# Load the YOLO model
model = YOLO(model_path)

classNames = ["Capacitor", "Diode", "Dot-Cut Mark", "Excess-Solder", "IC", "MCU", "Missing Com.", "Non-Good com.",
              "Resistor", "Short", "Soldering-Missing", "Tilt-Com"]

def predict(chosen_model, img, classes=[], conf=0.5):
    results = chosen_model.predict(img, conf=conf)
    
    if classes:
        filtered_results = []
        for result in results:
            filtered_boxes = [box for box in result.boxes if result.names[int(box.cls[0])] in classes]
            result.boxes = filtered_boxes
            filtered_results.append(result)
        return filtered_results
    else:
        return results

def predict_and_detect(chosen_model, img, classes=[], conf=0.5):
    img_copy = img.copy()  # Create a copy of the original image
    results = predict(chosen_model, img_copy, classes, conf)
    bounding_box_predictions = []  # Empty list to store bounding box predictions

    for result in results:
        for idx, box in enumerate(result.boxes):
            class_name = result.names[int(box.cls[0])]
            confidence = float(box.conf)
            x1, y1, x2, y2 = box.xyxy[0].numpy()  # Convert tensor to numpy
            bounding_box_predictions.append({"Label": class_name, "Confidence": confidence, "x1": x1, "y1": y1, "x2": x2, "y2": y2})

            if class_name in ["Capacitor", "Diode", "IC", "MCU", "Dot-Cut Mark", "Resistor"]:
                class_color = (0, 255, 0)  # Green color for certain classes
                Actual_Results = "OK"
            elif class_name in ["Excess-Solder", "Missing Com.", "Non-Good com.", "Short", "Soldering-Missing", "Tilt-Com"]:
                class_color = (0, 0, 255)  # Red color for certain classes
                Actual_Results = "FAIL"
            else:
                class_color = (255, 255, 255)  # Default color for other classes 
                Actual_Results = "UNKNOWN"
            
            # Convert the datatype of the image to uint8
            img_copy = img_copy.astype('uint8')
            
            cv2.rectangle(img_copy, (int(x1), int(y1)),
                          (int(x2), int(y2)), class_color, 2)
            cv2.putText(img_copy, f"{class_name}",
                        (int(x1), int(y1) - 10),
                        cv2.FONT_HERSHEY_PLAIN, 2, class_color, 2, cv2.LINE_AA)

    return img_copy, bounding_box_predictions

class YOLOTransformer(VideoTransformerBase):
    def __init__(self):
        self.model = model

    def transform(self, frame):
        img = frame.to_ndarray(format="bgr24")
        result_img, bounding_box_predictions = predict_and_detect(self.model, img, classes=selected_labels, conf=0.5)
        
        # Save bounding box predictions to Excel file
        for prediction in bounding_box_predictions:
            # Convert tensor values to Python native types (float or int)
            confidence = float(prediction["Confidence"]) * 100
            confidence_str = f"{confidence:.2f}%"
            x1 = int(prediction["x1"])
            y1 = int(prediction["y1"])
            x2 = int(prediction["x2"])
            y2 = int(prediction["y2"])
            Actual_Results = "UNKNOWN"
            
            # Determine success rate based on class
            if prediction["Label"] in ["Capacitor", "Diode", "IC", "MCU", "Dot-Cut Mark", "Resistor"]:
                Actual_Results = "OK"
                cell_color = "00FF00"  # Green color in HEX format
            elif prediction["Label"] in ["Excess-Solder", "Missing Com.", "Non-Good com.", "Short", "Soldering-Missing", "Tilt-Com"]:
                Actual_Results = "FAIL"
                cell_color = "FF0000"  # Red color in HEX format
            if 50 <= confidence < 90:
                Actual_Results = "NOT OK"
                cell_color = "FFFF00"  # Yellow color in HEX format
 
            # Determine Prediction Accuracy
            if 85 <= confidence <= 100:
                Prediction_Accuracy = "PASS"
                accuracy_color = "00FF00"  # Green color in HEX format
            elif 50 <= confidence < 85:
                Prediction_Accuracy = "FAIL"
                accuracy_color = "FF0000"  # Red color in HEX format
            else:
                Prediction_Accuracy = "UNKNOWN"
                accuracy_color = "FFFFFF"  # White color in HEX format

            # Create a tuple of the prediction to check for duplicates
            prediction_tuple = (prediction["Label"], confidence, x1, y1, x2, y2)
            
            # Only add unique predictions
            if prediction_tuple not in unique_predictions:
                unique_predictions.add(prediction_tuple)
                # Append the bounding box predictions to the Excel sheet
                ws.append([prediction["Label"], row_number, confidence_str, x1, y1, x2, y2, Actual_Results, Prediction_Accuracy])
                
                # Apply color formatting to the "success_rate" column
                row = ws.max_row
                ws.cell(row=row, column=8).fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
                ws.cell(row=row, column=9).fill = PatternFill(start_color=accuracy_color, end_color=accuracy_color, fill_type="solid")

                # Apply red color formatting to the "Label" column if the label is in the specified list
                if prediction["Label"] in ["Excess-Solder", "Missing Com.", "Non-Good com.", "Short", "Soldering-Missing", "Tilt-Com"]:
                    ws.cell(row=row, column=1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                row_number += 1  # Increment row number

        # Save the changes to the Excel file
        wb.save(filename)

        return result_img

# Streamlit interface
st.title("AOI Live Object Detection")

# Create a multiselect widget for selecting labels
selected_labels = st.sidebar.multiselect("Select Labels", classNames, default=classNames)

# Create a placeholder to display bounding box predictions in the sidebar
bounding_box_placeholder = st.sidebar.empty()

# Upload XLSX file
uploaded_file = st.file_uploader("Upload XLSX file", type=["xlsx"])

# Initialize the workbook and worksheet for storing results
filename = "results.xlsx"
try:
    if uploaded_file is not None:
        wb = load_workbook(uploaded_file)
        ws = wb.active
    else:
        wb = load_workbook(filename) if os.path.exists(filename) else Workbook()
        ws = wb.active
        # Add the custom column headers if the worksheet is empty
        if ws.max_row == 1:
            ws.append(["Label", "S.No", "Confidence", "x1", "y1", "x2", "y2", "Actual Results", "Prediction Accuracy"])
            
            # Apply bold font to headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
except Exception as e:
    st.error(f"Error accessing workbook: {e}")

# Create download button for Excel file
if st.button("Download Report results.xlsx"):
    try:
        wb.save(filename)
        with open(filename, "rb") as file:
            b64 = base64.b64encode(file.read()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="results.xlsx">Download results.xlsx</a>'
        st.markdown(href, unsafe_allow_html=True)
    except Exception as e:
        st.error(f"Error generating download link: {e}")

unique_predictions = set()  # Track unique predictions to avoid duplicates
row_number = ws.max_row + 1 if ws.max_row > 1 else 1  # Adjust the row number to start from the first row if empty

# Activate the YOLO video transformer with specified dimensions
webrtc_streamer(
    key="example",
    video_transformer_factory=YOLOTransformer,
    media_stream_constraints={
        "video": {
            "width": {"ideal": 800},
            "height": {"ideal": 600},
        },
        "audio": False,
    },
    video_html_attrs={
        "style": {"width": "800px", "height": "600px"},
    }
)
